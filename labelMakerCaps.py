# -*- coding: utf-8 -*-

"""
Conda environment is created to supply dependencies for working with XSLX files
Commented out are the procedures to perform in Anaconda PowerShell
"""

# Conda env is mapped to "C:\Users\10351517\Miniconda3\envs\spyder-envP01"
# Conda env was created in Anaconda Prompt 
# conda create -n spyder-envP01 -y
# conda install spyder-kernels openpyxl -y
# conda install spyder-kernels pandas -y    

import pandas as pd

"""
TrackingSheet is the excel file being used by Stability team
Sheet "2022 Sample Details" is 
Row 1 is skipped, row 2 is used for column ID's 
"""

#Global info or variables
TrackingSheet = pd.read_excel("~\Documents\Code\Projects\Stability\Stability setup\P01\Label maker\FY22 Stability Tracking Sheet.xlsx", sheet_name="2022 Sample Details", skiprows=1)
labelTemplate = pd.DataFrame(columns = ["1", "2", "3", "4", "5"])
labelsDF = pd.DataFrame(columns = ["1", "2", "3", "4", "5", "6"])
timePoints = ["4C", "25C 0.5yr", "25C 1yr","25C 1.5yr", "25C 2yr","25C 3yr","25C 4yr", "25C 5yr" ]
date = "start: 02/08/22"

"""
Create a template for labels
"""
for index in TrackingSheet.index:
    
    # print(TrackingSheet["Batch Number"][index], 
    #       TrackingSheet["Quarter"][index], 
    #       TrackingSheet["Target"][index], 
    #       TrackingSheet["Format"][index])
    
    target = str(TrackingSheet["Target"][index]).strip()
    clone = str(TrackingSheet["Clone"][index]).strip()
    col1 = target[0:8] + " " + clone[0:4]
    # target = TrackingSheet["Target"][index] + " " + str(TrackingSheet["Clone"][index])
    
    labelTemplate.loc[index] = [col1, 
                          TrackingSheet["Format"][index],
                          TrackingSheet["Batch Number"][index],
                          TrackingSheet["Sample Concentration (mg/ml)"][index],
                          date,]
                          
"""
Apply template to make labels implementing timepoints
"""
    
for index in labelTemplate.index:
    
    # print(labelTemplate["1"][index], 
    #       labelTemplate["2"][index],
    #       labelTemplate["3"][index],
    #       labelTemplate["4"][index],
    #       labelTemplate["5"][index])
    
    tempDF = pd.DataFrame(columns = ["1", "2", "3", "4", "5", "6"])
    
    if labelTemplate["2"][index] == "Ab-E":
        # print("TRUE")
        
        for i in range(5):
        
            
            tempDF.loc[i] = [labelTemplate["1"][index],
                             labelTemplate["2"][index],
                             labelTemplate["3"][index],
                             labelTemplate["4"][index],
                             labelTemplate["5"][index],
                             timePoints[i]]
            
            continue
        
    else:
        
        # print("FALSE")
        
        for i in range(8):
            
            tempDF.loc[i] = [labelTemplate["1"][index],
                             labelTemplate["2"][index],
                             labelTemplate["3"][index],
                             labelTemplate["4"][index],
                             labelTemplate["5"][index],
                             timePoints[i]]
            
            continue
        
    frames = [labelsDF, tempDF]
        
    labelsDF = pd.concat(frames, ignore_index = True, )
            
    tempDF = pd.DataFrame(columns = ["1", "2", "3", "4", "5", "6"])
        
    continue

# labelsDF.to_csv("labelsDF.csv")

"""

Output XSLX file
Implement openpyxl to write dataframe to Excel file

1. Figure out how to write to individual cells - complete
2. Figure out how to write XLSX file - complete
3. Figure out how to match cell to csv position and copy over values - pending
4. Add in column headers
5. Figure out how to format each cell to work with printer

"""

from openpyxl import Workbook

fileName = "~\Documents\Code\Projects\Stability\test.xlsx"
workbook = Workbook()
sheet = workbook.active
columnABC = ["A","B","C","D","E","F"]

"""

Making headers for file

"""

# for value in range (1,7):
    
#     colCells = sheet.cell(row = 1, column = value)
#     colCells.value = value

"""

Code for matching Excel cells to copy over DF 

"""

# for columnNum in range(0,6):
    
#     for i in range(0, 20):
        
#         # print(letter, i, sep="")
#         # i = str(i)
#         # cell = "{}{}".format(letter,i)
#         # print(cell)
#         # sheet[cell] = "nathan"
        
#         print(i,columnNum)
#         print(labelsDF.iloc[ (i), (columnNum) ])
        
        
#         mycell = sheet.cell(row = i+1, column = columnNum+1)
#         mycell.value = "To do still"
        

# workbook.save(filename=fileName)


"""
Using Panda dataframe.f(x) for excel file output
"""

labelsDF.to_excel(fileName, index=False, engine='openpyxl')


        
        
        
        